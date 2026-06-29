from datetime import datetime
from io import BytesIO
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


def _clean(value: str) -> str:
    """Убирает лишние пробелы и заменяет неразрывные пробелы на обычные."""
    return value.strip().replace("\xa0", " ")


def parse_employee_directory(file_path: str | bytes) -> list[dict]:
    """Парсит Excel-файл с данными сотрудников и возвращает список словарей с данными каждого сотрудника."""
    # Если file_path - это байты (например, из загруженного файла), то оборачиваем его в BytesIO, чтобы load_workbook мог его прочитать
    if isinstance(file_path, bytes):
        file_path = BytesIO(file_path)
    try:
        wb = load_workbook(filename=file_path, read_only=True)
    # Если файл повреждён или не является Excel-файлом, выбрасываем ValueError
    except (InvalidFileException, BadZipFile):
        raise ValueError("Файл повреждён или не является Excel-файлом")
    try:
        ws = wb.active  # Получаем активный лист

        employees = []
        header_found = False  # Флаг для определения, когда найдены данные сотрудников

        for row in ws.iter_rows(values_only=True): # Итерируемся по строкам листа, получаем только значения ячеек
            if row[0] == "Сотрудник":  # Ищем строку, где первый столбец == "Сотрудник", чтобы начать парсинг данных
                header_found = True  # Устанавливаем флаг, что заголовок найден, и следующие строки будут данными сотрудников
                continue

            if header_found:
                if row[0] is not None: # Если первый столбец не пустой, считаем, что это данные сотрудника
                    employee = {
                        # Убираем лишние пробелы и заменяем неразрывные пробелы на обычные, чтобы данные были чистыми
                        "full_name": _clean(row[0]),
                        "position": _clean(row[12]),
                        "contract_number": _clean(row[3]),
                        "contract_date": datetime.strptime(row[8], "%d.%m.%Y").date(), # Преобразуем строку с датой в объект date, предполагая формат "день.месяц.год"
                        "document_type": _clean(row[13]),
                        "id_series": _clean(row[14]),
                        "id_number": _clean(row[15]),
                        "id_issued_date": datetime.strptime(row[16], "%d.%m.%Y").date(),
                        "issued_by": _clean(row[17]),
                        "address": _clean(row[18])
                    }
                    employees.append(employee)
                else:  # Иначе, считаем, что данные сотрудников закончились
                    break
    finally:
        wb.close()
    if not employees:
        raise ValueError("В файле не найдены необходимые данные. Проверьте, что загружен правильный файл.")
    return employees



def employee_name(value: str) -> bool:
    """
    Проверяет, является ли значение ФИО сотрудника.
    Если строка содержит 3 слова, каждое из которых начинается с заглавной буквы, то это имя сотрудника.
    """
    parts = value.split()
    result = len(parts) == 3 and all(p and p[0].isupper() for p in parts)
    return result


def inventory_code(value: str) -> bool:
    """
    Проверяет, является ли значение кодом инвентаря.
    """
    return value.startswith("ЦБ") or value.startswith("БШ")


def parse_inventory(file_path: str | bytes) -> list[dict]:
    """Парсит Excel-файл с данными инвентаря и возвращает список словарей с данными каждого предмета."""
    if isinstance(file_path, bytes):
        file_path = BytesIO(file_path)
    try:
        wb = load_workbook(filename=file_path, read_only=True)
    # Если файл повреждён или не является Excel-файлом, выбрасываем ValueError
    except (InvalidFileException, BadZipFile):
        raise ValueError("Файл повреждён или не является Excel-файлом")
    try:
        ws = wb.active

        current_employee = None  # Храним текущего сотрудника
        current_tool_code = None  # Храним текущий код инструмента
        current_tool_name = None  # Храним текущее название инструмента
        price = None  # Храним цену текущего инструмента
        inventories = []

        for row in ws.iter_rows(values_only=True):
            if row[0] is None or row[1] == "Кол.":
                if current_tool_code and price is not None:
                    quantity = row[4]
                    inventories.append({
                        "employee_name": current_employee,
                        "tool_name": current_tool_name,
                        "tool_code": current_tool_code,
                        "quantity": quantity,
                        "price": price,
                    })
                    current_tool_code = None  # Сбрасываем
                    price = None
                continue

            # Убираем лишние пробелы и заменяем неразрывные пробелы на обычные, чтобы данные были чистыми
            value = str(row[0]).strip().replace("\xa0", " ")

            if value == "МЦ.04":
                continue

            # Если встречаем ФИО сотрудника
            elif employee_name(value):
                current_employee = value

            # Если встречаем код инвентаря
            elif inventory_code(value):
                current_tool_code = value
                price = row[4]

            # Если есть ФИО сотрудника, то следующая строка - это название инструмента
            elif current_employee:
                current_tool_name = value
    finally:
        wb.close()
    if not inventories:
        raise ValueError("В файле не найдены необходимые данные. Проверьте, что загружен правильный файл.")
    return inventories